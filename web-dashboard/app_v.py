import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================================
# Product Data Generation
# ============================================================================

def generate_products_from_sample():
    """Generate comprehensive product data based on sample"""
    
    # Base products from your sample
    base_products = [
        {
            'product_number': 'BBS-1001',
            'name': 'BBS Modem Alpha',
            'model': 'BBS-1001',
            'created_at': '2026-03-11T18:54:13.565786',
            'category': 'Electronics',
            'price': 299.99,
            'quantity': 50,
            'unit': 'pcs',
            'supplier': 'BBS Technologies',
            'min_quantity': 10,
            'description': 'High-speed modem with advanced features'
        },
        {
            'product_number': 'AVS-2001',
            'name': 'AVS Modem Pro',
            'model': 'AVS-2001',
            'created_at': '2026-03-11T18:54:13.581598',
            'category': 'Electronics',
            'price': 399.99,
            'quantity': 35,
            'unit': 'pcs',
            'supplier': 'AVS Systems',
            'min_quantity': 8,
            'description': 'Professional grade modem with enhanced performance'
        },
        {
            'product_number': 'PREF-001',
            'name': 'App DBBS',
            'model': 'DBBS',
            'created_at': '2026-03-11T18:55:10.963155',
            'category': 'Software',
            'price': 199.99,
            'quantity': 100,
            'unit': 'licenses',
            'supplier': 'DBBS Software',
            'min_quantity': 20,
            'description': 'Database management software'
        }
    ]
    
    # Additional products to populate the database
    additional_products = [
        {
            'product_number': 'NET-3001',
            'name': 'Network Router X9',
            'model': 'NR-X9',
            'created_at': datetime.now().isoformat(),
            'category': 'Networking',
            'price': 599.99,
            'quantity': 25,
            'unit': 'pcs',
            'supplier': 'Network Solutions Inc',
            'min_quantity': 5,
            'description': 'Gigabit router with advanced security features'
        },
        {
            'product_number': 'CAB-4001',
            'name': 'HDMI Cable 2.1',
            'model': 'HDMI-21',
            'created_at': datetime.now().isoformat(),
            'category': 'Accessories',
            'price': 29.99,
            'quantity': 500,
            'unit': 'pcs',
            'supplier': 'CableMaster',
            'min_quantity': 50,
            'description': 'High-speed HDMI 2.1 cable, 6ft'
        },
        {
            'product_number': 'MON-5001',
            'name': '4K Monitor Ultra',
            'model': '4K-Ultra',
            'created_at': datetime.now().isoformat(),
            'category': 'Displays',
            'price': 449.99,
            'quantity': 15,
            'unit': 'pcs',
            'supplier': 'DisplayTech',
            'min_quantity': 3,
            'description': '27-inch 4K UHD monitor with HDR'
        },
        {
            'product_number': 'KEY-6001',
            'name': 'Mechanical Keyboard',
            'model': 'MK-Pro',
            'created_at': datetime.now().isoformat(),
            'category': 'Peripherals',
            'price': 129.99,
            'quantity': 75,
            'unit': 'pcs',
            'supplier': 'KeyMaster',
            'min_quantity': 15,
            'description': 'RGB mechanical keyboard with Cherry MX switches'
        },
        {
            'product_number': 'MOU-7001',
            'name': 'Wireless Mouse',
            'model': 'WM-Gaming',
            'created_at': datetime.now().isoformat(),
            'category': 'Peripherals',
            'price': 59.99,
            'quantity': 120,
            'unit': 'pcs',
            'supplier': 'MouseTech',
            'min_quantity': 20,
            'description': 'High-precision wireless gaming mouse'
        },
        {
            'product_number': 'SSD-8001',
            'name': '1TB NVMe SSD',
            'model': 'NVMe-1TB',
            'created_at': datetime.now().isoformat(),
            'category': 'Storage',
            'price': 149.99,
            'quantity': 40,
            'unit': 'pcs',
            'supplier': 'StoragePro',
            'min_quantity': 10,
            'description': 'Ultra-fast NVMe SSD for gaming and professional use'
        },
        {
            'product_number': 'RAM-9001',
            'name': '16GB DDR4 RAM',
            'model': 'DDR4-16G',
            'created_at': datetime.now().isoformat(),
            'category': 'Components',
            'price': 89.99,
            'quantity': 60,
            'unit': 'pcs',
            'supplier': 'MemoryExperts',
            'min_quantity': 15,
            'description': 'High-performance DDR4 RAM 3200MHz'
        },
        {
            'product_number': 'CPU-10001',
            'name': 'Intel i7 Processor',
            'model': 'i7-12700K',
            'created_at': datetime.now().isoformat(),
            'category': 'Components',
            'price': 389.99,
            'quantity': 20,
            'unit': 'pcs',
            'supplier': 'Intel Corp',
            'min_quantity': 5,
            'description': '12th Gen Intel Core i7 processor'
        },
        {
            'product_number': 'GPU-11001',
            'name': 'RTX 4070 Graphics Card',
            'model': 'RTX-4070',
            'created_at': datetime.now().isoformat(),
            'category': 'Components',
            'price': 599.99,
            'quantity': 10,
            'unit': 'pcs',
            'supplier': 'NVIDIA',
            'min_quantity': 2,
            'description': 'High-performance gaming graphics card'
        },
        {
            'product_number': 'PWR-12001',
            'name': '750W Power Supply',
            'model': 'PSU-750',
            'created_at': datetime.now().isoformat(),
            'category': 'Components',
            'price': 119.99,
            'quantity': 30,
            'unit': 'pcs',
            'supplier': 'PowerTech',
            'min_quantity': 8,
            'description': 'Gold certified modular power supply'
        }
    ]
    
    # Combine all products
    all_products = base_products + additional_products
    
    # Create DataFrame
    products_df = pd.DataFrame(all_products)
    
    # Add additional fields
    products_df['status'] = 'Active'
    products_df['warranty_months'] = [24, 24, 12, 36, 12, 24, 12, 12, 24, 36, 24, 24, 36]
    products_df['weight_kg'] = [0.5, 0.6, 0.0, 1.2, 0.2, 0.8, 0.1, 0.1, 0.05, 0.8, 0.9, 1.1, 0.8]
    products_df['manufacturer'] = ['BBS', 'AVS', 'DBBS', 'NetGear', 'CableMaster', 'Samsung', 
                                   'Logitech', 'Razer', 'Western Digital', 'Corsair', 'Intel', 'NVIDIA', 'EVGA']
    
    return products_df

def generate_categories():
    """Generate categories data"""
    categories = [
        {'name': 'Electronics', 'description': 'Electronic devices and components'},
        {'name': 'Software', 'description': 'Software applications and licenses'},
        {'name': 'Networking', 'description': 'Network equipment and accessories'},
        {'name': 'Accessories', 'description': 'Various accessories and peripherals'},
        {'name': 'Displays', 'description': 'Monitors and display devices'},
        {'name': 'Peripherals', 'description': 'Input devices and peripherals'},
        {'name': 'Storage', 'description': 'Storage devices and media'},
        {'name': 'Components', 'description': 'Computer components and parts'}
    ]
    
    return pd.DataFrame(categories)

def generate_stock_movements(products_df, num_movements=50):
    """Generate stock movement history"""
    movements = []
    movement_types = ['INCREASE', 'DECREASE']
    reasons = [
        'New stock received',
        'Customer purchase',
        'Stock adjustment',
        'Return from customer',
        'Damaged goods',
        'Quality control',
        'Inventory count adjustment',
        'Transfer from warehouse',
        'Promotional sale',
        'Bulk order'
    ]
    
    start_date = datetime(2026, 1, 1)
    end_date = datetime(2026, 3, 24)
    
    for i in range(num_movements):
        product = products_df.sample(1).iloc[0]
        movement_type = random.choice(movement_types)
        quantity = random.randint(1, 50)
        reason = random.choice(reasons)
        
        # Generate random date between start and end
        random_date = start_date + timedelta(
            seconds=random.randint(0, int((end_date - start_date).total_seconds()))
        )
        
        # Calculate previous and new quantities
        previous_quantity = random.randint(0, 200)
        if movement_type == 'INCREASE':
            new_quantity = previous_quantity + quantity
        else:
            new_quantity = max(0, previous_quantity - quantity)
            quantity = abs(previous_quantity - new_quantity)
        
        movements.append({
            'product_id': i % len(products_df) + 1,
            'product_number': product['product_number'],
            'product_name': product['name'],
            'movement_type': movement_type,
            'quantity': quantity,
            'previous_quantity': previous_quantity,
            'new_quantity': new_quantity,
            'reason': reason,
            'created_at': random_date.isoformat()
        })
    
    return pd.DataFrame(movements)

def generate_suppliers():
    """Generate supplier data"""
    suppliers = [
        {'name': 'BBS Technologies', 'contact': 'contact@bbstech.com', 'phone': '+1-555-0101', 'address': '123 Tech Street, Silicon Valley, CA'},
        {'name': 'AVS Systems', 'contact': 'sales@avs.com', 'phone': '+1-555-0102', 'address': '456 Innovation Ave, Austin, TX'},
        {'name': 'DBBS Software', 'contact': 'info@dbbs.com', 'phone': '+1-555-0103', 'address': '789 Code Lane, Seattle, WA'},
        {'name': 'Network Solutions Inc', 'contact': 'support@netsol.com', 'phone': '+1-555-0104', 'address': '321 Connectivity Blvd, Boston, MA'},
        {'name': 'CableMaster', 'contact': 'orders@cablemaster.com', 'phone': '+1-555-0105', 'address': '654 Wire Way, Chicago, IL'},
        {'name': 'DisplayTech', 'contact': 'sales@displaytech.com', 'phone': '+1-555-0106', 'address': '987 Pixel Drive, New York, NY'},
        {'name': 'KeyMaster', 'contact': 'info@keymaster.com', 'phone': '+1-555-0107', 'address': '147 Click Street, Denver, CO'},
        {'name': 'MouseTech', 'contact': 'support@mousetech.com', 'phone': '+1-555-0108', 'address': '258 Scroll Ave, Portland, OR'},
        {'name': 'StoragePro', 'contact': 'sales@storagepro.com', 'phone': '+1-555-0109', 'address': '369 Data Drive, San Jose, CA'},
        {'name': 'MemoryExperts', 'contact': 'info@memoryexperts.com', 'phone': '+1-555-0110', 'address': '741 RAM Road, Phoenix, AZ'},
        {'name': 'Intel Corp', 'contact': 'sales@intel.com', 'phone': '+1-555-0111', 'address': '852 Processor Way, Santa Clara, CA'},
        {'name': 'NVIDIA', 'contact': 'info@nvidia.com', 'phone': '+1-555-0112', 'address': '963 Graphics Blvd, Santa Clara, CA'},
        {'name': 'PowerTech', 'contact': 'support@powertech.com', 'phone': '+1-555-0113', 'address': '159 Watt Street, Dallas, TX'}
    ]
    
    return pd.DataFrame(suppliers)

def create_styled_excel(filename='product_data.xlsx'):
    """Create a styled Excel file with multiple sheets"""
    
    print("📊 Generating product data...")
    products_df = generate_products_from_sample()
    categories_df = generate_categories()
    suppliers_df = generate_suppliers()
    stock_movements_df = generate_stock_movements(products_df, 30)
    
    # Create summary statistics
    summary_data = {
        'Metric': [
            'Total Products',
            'Total Categories',
            'Total Suppliers',
            'Total Stock Value',
            'Average Product Price',
            'Total Units in Stock',
            'Low Stock Items',
            'Out of Stock Items'
        ],
        'Value': [
            len(products_df),
            len(categories_df),
            len(suppliers_df),
            f"${(products_df['price'] * products_df['quantity']).sum():,.2f}",
            f"${products_df['price'].mean():,.2f}",
            products_df['quantity'].sum(),
            len(products_df[products_df['quantity'] <= products_df['min_quantity']]),
            len(products_df[products_df['quantity'] == 0])
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    
    # Create Excel file with styling
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write DataFrames to sheets
        products_df.to_excel(writer, sheet_name='Products', index=False)
        categories_df.to_excel(writer, sheet_name='Categories', index=False)
        suppliers_df.to_excel(writer, sheet_name='Suppliers', index=False)
        stock_movements_df.to_excel(writer, sheet_name='Stock Movements', index=False)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Get workbook and worksheets
        workbook = writer.book
        
        # Style each sheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Style header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
    
    print(f"✅ Excel file created successfully: {filename}")
    return filename

def create_import_template():
    """Create a template for importing products"""
    
    # Create template DataFrame with example data
    template_data = [
        {
            'product_number': 'NEW-001',
            'name': 'New Product Example',
            'category': 'Electronics',
            'description': 'This is an example product for import',
            'price': 99.99,
            'quantity': 100,
            'unit': 'pcs',
            'supplier': 'Example Supplier',
            'min_quantity': 10
        },
        {
            'product_number': 'NEW-002',
            'name': 'Another Product',
            'category': 'Software',
            'description': 'Software license example',
            'price': 199.99,
            'quantity': 50,
            'unit': 'licenses',
            'supplier': 'Software Supplier',
            'min_quantity': 5
        }
    ]
    
    template_df = pd.DataFrame(template_data)
    
    # Create categories template
    categories_template = pd.DataFrame([
        {'name': 'New Category 1', 'description': 'Description for category 1'},
        {'name': 'New Category 2', 'description': 'Description for category 2'}
    ])
    
    # Create Excel file
    filename = 'product_import_template.xlsx'
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        template_df.to_excel(writer, sheet_name='Products', index=False)
        categories_template.to_excel(writer, sheet_name='Categories', index=False)
        
        # Add instructions sheet
        instructions = pd.DataFrame({
            'Step': ['1', '2', '3', '4', '5'],
            'Instruction': [
                'Fill in your product data in the Products sheet',
                'Fill in categories in the Categories sheet (optional)',
                'Product number must be unique',
                'Category names will be created automatically if they don\'t exist',
                'Save the file and upload to the system'
            ]
        })
        instructions.to_excel(writer, sheet_name='Instructions', index=False)
        
        # Style the template
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Style header
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✅ Import template created: {filename}")
    return filename

def create_sample_data_for_demo():
    """Create a comprehensive sample data file for demonstration"""
    
    print("🚀 Creating comprehensive sample data for Product Management System...")
    
    # Create main product data
    products_df = generate_products_from_sample()
    
    # Add some specific products from your sample
    sample_products = [
        {
            'product_number': 'BBS-1001',
            'name': 'BBS Modem Alpha',
            'category': 'Electronics',
            'description': 'High-performance modem with advanced features',
            'price': 299.99,
            'quantity': 50,
            'unit': 'pcs',
            'supplier': 'BBS Technologies',
            'min_quantity': 10,
            'status': 'Active',
            'warranty_months': 24
        },
        {
            'product_number': 'AVS-2001',
            'name': 'AVS Modem Pro',
            'category': 'Electronics',
            'description': 'Professional grade modem with enhanced performance',
            'price': 399.99,
            'quantity': 35,
            'unit': 'pcs',
            'supplier': 'AVS Systems',
            'min_quantity': 8,
            'status': 'Active',
            'warranty_months': 24
        },
        {
            'product_number': 'PREF-001',
            'name': 'App DBBS',
            'category': 'Software',
            'description': 'Database management software',
            'price': 199.99,
            'quantity': 100,
            'unit': 'licenses',
            'supplier': 'DBBS Software',
            'min_quantity': 20,
            'status': 'Active',
            'warranty_months': 12
        }
    ]
    
    # Create final products DataFrame
    final_products_df = pd.DataFrame(sample_products + [
        {
            'product_number': f'PRD-{i:04d}',
            'name': f'Product {i}',
            'category': np.random.choice(['Electronics', 'Software', 'Accessories', 'Networking']),
            'description': f'Description for product {i}',
            'price': np.random.uniform(10, 500),
            'quantity': np.random.randint(0, 200),
            'unit': np.random.choice(['pcs', 'kg', 'liters', 'boxes']),
            'supplier': np.random.choice(['Supplier A', 'Supplier B', 'Supplier C']),
            'min_quantity': np.random.randint(5, 50),
            'status': 'Active',
            'warranty_months': np.random.choice([12, 24, 36])
        }
        for i in range(20, 30)
    ])
    
    # Create Excel file
    filename = 'demo_product_data.xlsx'
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        final_products_df.to_excel(writer, sheet_name='Products', index=False)
        generate_categories().to_excel(writer, sheet_name='Categories', index=False)
        generate_suppliers().to_excel(writer, sheet_name='Suppliers', index=False)
        
        # Add summary
        summary_df = pd.DataFrame({
            'Metric': ['Total Products', 'Total Value', 'Average Price', 'Total Quantity'],
            'Value': [
                len(final_products_df),
                f"${(final_products_df['price'] * final_products_df['quantity']).sum():,.2f}",
                f"${final_products_df['price'].mean():,.2f}",
                final_products_df['quantity'].sum()
            ]
        })
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    print(f"✅ Demo data created: {filename}")
    return filename

# ============================================================================
# Main execution
# ============================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("📦 Product Management System - Excel Data Generator")
    print("=" * 60)
    
    # Create all required Excel files
    create_styled_excel('product_data.xlsx')
    create_import_template()
    create_sample_data_for_demo()
    
    print("\n" + "=" * 60)
    print("✅ All Excel files created successfully!")
    print("\nFiles created:")
    print("  1. product_data.xlsx - Full product database with styling")
    print("  2. product_import_template.xlsx - Template for importing new products")
    print("  3. demo_product_data.xlsx - Sample data for demonstration")
    print("\n📌 You can now upload these files to the Product Management System")
    print("=" * 60)
    